import subprocess, sys
subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q', '--break-system-packages'], capture_output=True)

from openpyxl import load_workbook
import json, os

path = r'C:\Amalgamate\Projects\Zawadi SMS\data\expenses vote heads.xlsx'
wb = load_workbook(path, read_only=True)
print("SHEETS:", wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            print(list(row))
