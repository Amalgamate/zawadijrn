import subprocess, sys, json, os

subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q', '--break-system-packages'], capture_output=True)
from openpyxl import load_workbook

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'expenses vote heads.xlsx')
print("Reading:", path)

wb = load_workbook(path, read_only=True)
print("Sheets:", wb.sheetnames)

result = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            rows.append([str(v) if v is not None else '' for v in row])
    result[sheet_name] = rows

output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'expenses_extracted.json')
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)

print("Done. Written to:", output_path)
for sheet, rows in result.items():
    print(f"\n=== {sheet} ({len(rows)} rows) ===")
    for r in rows:
        print(r)
